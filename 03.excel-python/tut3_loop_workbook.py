from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('tim1.xlsx')
ws = wb.active

for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        cells = ws[char + str(row)]
        # this will change the value to the name of the cel
        ws[char + str(row)] = char + str(row)
        print('cel:', cells)
        print('cel value:', cells.value)


wb.save('tim2.xlsx')
