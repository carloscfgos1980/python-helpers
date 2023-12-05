from openpyxl import Workbook, load_workbook

wb = load_workbook(r'/Users/carlosinfante/Desktop/subjetcs-data.xlsx')
ws = wb.active  # give the actual worksheet
wb.create_sheet('test1')  # create a new worksheet
wb.create_sheet('test2')

print('worksheet:', ws)
print('A2 value:', ws['A2'].value)
print('wb.sheetnames:', wb.sheetnames)
print(wb['test'])

ws['A2'].value = 'Carlos'
ws['A3'].value = 'Jose'
ws.merge_cells("A1:D1")  # Merge cells method
ws.unmerge_cells("A1:D1")  # Unmerge cells method

ws.insert_rows(7)  # insert an empty row
ws.delete_rows(7)  # delete row
ws.insert_cols(2)  # insert a column at the second column (B)
ws.delete_cols(2)  # delete the second column (B)

# This moves the selected cells 2 rows bellow and 2 rows to the right. If we put minus sign then we get the contrry effect
ws.move_range("A1:D10", rows=2, cols=2)

wb.save(r'/Users/carlosinfante/Desktop/subjetcs-data.xlsx')
wb.save('subjetcs-data1.xlsx')

print(ws['A3'].value)
