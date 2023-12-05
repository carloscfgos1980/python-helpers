import openpyxl

# create a new workbook
new_book = openpyxl.Workbook()

# get a reference to the active sheet
current_sheet = new_book.active

# write informatin to the cells

current_sheet["A1"].value = "The meaning of life"
current_sheet["A2"] = "Carlos"  # no need of .value with this method
# (1,2) => 1 represent the row, 2 represent the column
current_sheet.cell(1, 2).value = 42
# diferent way to do it, the info is the 3rd value in the tupple
current_sheet.cell(2, 2, 43)

# save this file
new_book.save("tutorial1.xlsx")
