import openpyxl

# create a new workbook
new_workbook = openpyxl.Workbook()  # will automatically have a sheet

# Add a sheet at the last position
musical_sheet = new_workbook.create_sheet("Musicals")

# Now we have 2 sheets: Sheet / Musicals

# Add a worksheet at the start
animation_sheet = new_workbook.create_sheet("Animations", 0)

# Now we have 3 sheets: Animations / Sheet / Musicals

# Add a worksheet at the penultime position
comedy_sheet = new_workbook.create_sheet("Comedies", -1)

# Now we have 4 sheets: Animations / Sheet / Animations / Comedies / Musicals

# change the name of workbook "Sheet"
# [1] is because is worksheet "Sheet" is on the 2nd column
new_workbook.worksheets[1].title = "Dramas"

# Add some info in each sheet
animation_sheet["A1"] = "Up"
comedy_sheet["A1"] = "Borat"
# this one we didn't create a worksheet with the name. That is why we need to reference to the workbook
new_workbook["Dramas"]["A1"] = "Gone with the wind"
musical_sheet["A1"] = "La la land"


new_workbook.save("movies.xlsx")
new_workbook.close()
